# -*- coding: utf-8 -*-
from tkinter import filedialog, Tk
import aspose.pdf as pdf
import io
import csv
import openpyxl
import gc
import pandas as pd
import re
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.styles import Font
import numpy as np



def main():
    # Initialiser la fenêtre Tkinter pour la sélection des fichiers
    root = Tk()
    root.title("Sélectionnez des fichiers PDF")
    root.geometry("300x200")

    file_paths = filedialog.askopenfilenames(
        title="Sélectionnez un fichier PDF",
        filetypes=[("Fichiers PDF", "*.pdf")]
    )
    if not file_paths:
        print("Aucun fichier sélectionné. Fin du programme.")
        root.destroy()
        root.quit() 
        return

    # Sélectionner le chemin de sauvegarde pour le CSV combiné
    csv_output_path = filedialog.asksaveasfilename(
        title="Enregistrer le fichier CSV combiné",
        defaultextension=".csv",
        filetypes=[("Fichiers CSV", "*.csv")]
    )
    if not csv_output_path:
        print("Aucun chemin de sauvegarde sélectionné. Fin du programme.")
        root.destroy()
        root.quit()
        return

    process_pdfs_to_csv(file_paths, csv_output_path)
    print(f"Fichier CSV combiné créé avec succès : {csv_output_path}")

    # Convertir le fichier CSV en liste
    liste_donnees = csv_to_list(csv_output_path)
    if liste_donnees:
        # Créer un dictionnaire à partir des listes
        dictionnaire = create_dictionaries_from_lists(liste_donnees)

        # Nettoyer le dictionnaire
        dictionnaire_nettoye = nettoyer_dictionnaire(dictionnaire)

        # Afficher le dictionnaire nettoyé
        print("Dictionnaire nettoyé :")
        print(dictionnaire_nettoye)

        # Appeler la fonction pour détecter et traiter les informations
        detecter_et_traiter_informations(dictionnaire_nettoye)


def process_pdfs_to_csv(file_paths, csv_output_path):
    try:
        with open(csv_output_path, mode='w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)

            # Parcourir tous les fichiers PDF
            for file_path in file_paths:
                try:
                    print(f"Traitement du fichier : {file_path}")
                    document = pdf.Document(file_path)

                    # Créer un nouveau document contenant uniquement la première page
                    first_page_document = pdf.Document()
                    first_page_document.pages.add(document.pages[1])

                    excel_save_options = pdf.ExcelSaveOptions()
                    excel_save_options.format = pdf.ExcelSaveOptions.ExcelFormat.XLSX

                    output_stream = io.BytesIO()

                    try:
                        # Sauvegarder la première page en mémoire temporaire
                        first_page_document.save(output_stream, excel_save_options)
                        output_stream.seek(0)

                        # Charger les données Excel avec openpyxl
                        workbook = openpyxl.load_workbook(output_stream)
                        worksheet = workbook.active

                        # Écriture progressive des données dans le CSV
                        for row in worksheet.iter_rows(values_only=True):
                            writer.writerow(row)

                    finally:
                        # Libérer les ressources
                        output_stream.close()
                        del output_stream
                        del workbook
                        del worksheet
                        del first_page_document
                        gc.collect()

                except Exception as e:
                    print(f"Erreur lors du traitement du fichier {file_path}: {e}")
                    continue
    except Exception as e:
        print(f"Erreur inattendue lors du traitement des fichiers PDF : {e}")


def csv_to_list(nom_fichier):
    liste_donnees = []
    try:
        with open(nom_fichier, mode='r', encoding='utf-8') as fichier_csv:
            lecteur_csv = csv.reader(fichier_csv)
            for ligne in lecteur_csv:
                liste_donnees.append(ligne)
    except Exception as e:
        print(f"Erreur lors de la conversion du CSV en liste : {e}")
    return liste_donnees


def create_dictionaries_from_lists(listes):
    dictionnaire = {}
    for i, liste in enumerate(listes):
        dictionnaire[f"Liste_{i+1}"] = liste
    return dictionnaire


def nettoyer_dictionnaire(resultats):
    dictionnaire_nettoye = {}
    for cle, liste in resultats.items():
        dictionnaire_nettoye[cle] = [str(element).replace("'", "").replace(",", "").replace(" ", "") for element in liste if element]
    return dictionnaire_nettoye

def detecter_et_traiter_informations(dictionnaire):
    resultats = {"Elong.4D": [], "Elong.5D": [], "InitialD": [], "Proof(0.2%)": [],"mE": [],"RT UTS": [],"450°C UTS":[],"BAR": [],"DIAMETER":[], "ElongatFracture": [],"ElongafterFracture": [] , "RT 0.2%Proof": [],"450°C 0.2%Proof": [], "HRC": [],"Moyenne_HRC": []}
    compteur_0_2_proof = 0
    compteur_UTS = 0
    listes_dictionnaire = list(dictionnaire.keys())
   
    for liste, contenu in dictionnaire.items():
        position_liste = listes_dictionnaire.index(liste) if liste in listes_dictionnaire else None  

        for index, element in enumerate(contenu):
        
            if "Elong.4D" in str(element):
    
                valeurs = []
                if position_liste is not None and position_liste + 2 < len(listes_dictionnaire):
                    deuxieme_liste = listes_dictionnaire[position_liste + 2]
                    contenu_deuxieme_liste = dictionnaire.get(deuxieme_liste, [])

                    if len(contenu_deuxieme_liste) > 4:
                        valeurs.append(contenu_deuxieme_liste[4])

                resultats["Elong.4D"].append(
                    # "Liste actuelle": liste,
                    # "Index détecté": index,
                     valeurs
                )

            if "Elong.5D" in str(element):
                

                valeurs = []
                if position_liste is not None and position_liste + 2 < len(listes_dictionnaire):
                    deuxieme_liste = listes_dictionnaire[position_liste + 2]
                    contenu_deuxieme_liste = dictionnaire.get(deuxieme_liste, [])
                    if len(contenu_deuxieme_liste) > 6:
                        valeurs.append(contenu_deuxieme_liste[6])

                resultats["Elong.5D"].append(
                    # "Liste actuelle": liste,
                    # "Index détecté": index,
                    valeurs
                )

            if "InitialD" in str(element):
                

                valeurs = []
                if position_liste is not None and position_liste + 2 < len(listes_dictionnaire):
                    deuxieme_liste = listes_dictionnaire[position_liste + 2]
                    contenu_deuxieme_liste = dictionnaire.get(deuxieme_liste, [])
                    if len(contenu_deuxieme_liste) > 7:
                        valeurs.append(contenu_deuxieme_liste[7])

                resultats["InitialD"].append(
                    # "Liste actuelle": liste,
                    # "Index détecté": index,
                    valeurs
                )
            
            if "Proof(0.2%)" in str(element):
         

             valeurs = []
                 
             if position_liste is not None and position_liste + 1 < len(listes_dictionnaire):
                     deuxieme_liste = listes_dictionnaire[position_liste + 1]
                     contenu_deuxieme_liste = dictionnaire.get(deuxieme_liste, [])

                     if len(contenu_deuxieme_liste) > 2:
                         valeurs.append(contenu_deuxieme_liste[2])

             resultats["Proof(0.2%)"].append(
             
              valeurs
              )

            if "mE" in str(element):
                valeurs = []
                if position_liste is not None and position_liste + 2 < len(listes_dictionnaire):
                    deuxieme_liste_suivante = listes_dictionnaire[position_liste + 2]
                    contenu_deuxieme_liste = dictionnaire.get(deuxieme_liste_suivante, [])

                    if index + 1 < len(contenu_deuxieme_liste):
                        valeurs.append(contenu_deuxieme_liste[index + 1])
                    else:
                        valeurs.append("Pas de valeur disponible")

                resultats["mE"].append(
                    
                     valeurs
                )
            
            if element == "UTS":
                compteur_UTS += 1
                valeurs = []
                if compteur_UTS >4:
                    compteur_UTS = 1

                if position_liste is not None and position_liste + 2 < len(listes_dictionnaire):
                    deuxieme_liste_suivante = listes_dictionnaire[position_liste + 2]
                    contenu_deuxieme_liste = dictionnaire.get(deuxieme_liste_suivante, [])

                    if compteur_UTS <= 2 and len(contenu_deuxieme_liste) > index + 5:
                        valeurs.append(contenu_deuxieme_liste[index + 5])
                    elif len(contenu_deuxieme_liste) > index + 4:
                        valeurs.append(contenu_deuxieme_liste[index + 4])
                    else:
                        valeurs.append("RIEN")

               
                resultats["RT UTS"].append(
                  
                    valeurs
                ) if compteur_UTS <= 2 else  resultats["450°C UTS"].append(
                    
                    valeurs
                )
            if "IDENTITY" in str(element):
                if index + 1 < len(contenu):
                    serial_match = re.split(r"SerialNo.|SERIALNo:|Serialno.|SerialNo.:", contenu[index + 1])

                    if len(serial_match) > 1:
                        resultats["BAR"].append(serial_match[0])  # Partie avant "SerialNo"
                        resultats["BAR"].append([""])
                        resultats["DIAMETER"].append(serial_match[1])  # Partie après "SerialNo"
                        resultats["DIAMETER"].append([""])
                    else:
                        resultats["BAR"].append(serial_match[0])  # Partie avant "SerialNo"
                        resultats["BAR"].append([""])

                    
                    
            if "ElongatFracture" in str(element):
                valeurs = []
    
                if position_liste is not None and position_liste + 2 < len(listes_dictionnaire):
                    deuxieme_liste = listes_dictionnaire[position_liste + 2]
                    contenu_deuxieme_liste = dictionnaire.get(deuxieme_liste, [])

                    if index + 5 < len(contenu_deuxieme_liste):
                        valeur_brute = contenu_deuxieme_liste[index + 5]
            
                # Extraction de la partie contenant une virgule et des chiffres avant le %
                        match = re.search(r"(\d+,\d+|\d+)%", valeur_brute)
                        if match:
                            
                            resultats["ElongatFracture"].append(
                   
                       match.group(1)
                       )     
            if "ElongafterFracture" in str(element):
                valeurs = [] 

                if position_liste is not None and position_liste + 2 < len(listes_dictionnaire):
                     deuxieme_liste = listes_dictionnaire[position_liste + 2]
                     contenu_deuxieme_liste = dictionnaire.get(deuxieme_liste, [])

                     if index + 5 < len(contenu_deuxieme_liste):
                         valeur_brute = contenu_deuxieme_liste[index + 5]
             
                 # Extraction de la partie contenant une virgule et des chiffres avant le %
                         match = re.search(r"(\d+,\d+|\d+)%", valeur_brute)
                         if match:
                             
                             resultats["ElongafterFracture"].append(
                    
                        match.group(1)
                        )
            
            if "0.2%Proof" in str(element):
                compteur_0_2_proof += 1
                valeurs = []
                if compteur_0_2_proof >4:
                    compteur_0_2_proof = 1

                if position_liste is not None and position_liste + 2 < len(listes_dictionnaire):
                    deuxieme_liste_suivante = listes_dictionnaire[position_liste + 2]
                    contenu_deuxieme_liste = dictionnaire.get(deuxieme_liste_suivante, [])

                    if compteur_0_2_proof <= 2 and len(contenu_deuxieme_liste) > index + 5:
                        valeurs.append(contenu_deuxieme_liste[index + 5])
                    elif len(contenu_deuxieme_liste) > index + 4:
                        valeurs.append(contenu_deuxieme_liste[index + 4])
                    else:
                        valeurs.append("RIEN")

                resultats["RT 0.2%Proof"].append(
                  
                    valeurs
                ) if compteur_0_2_proof <= 2 else  resultats["450°C 0.2%Proof"].append(
                    
                    valeurs
                )
            
            
            
            
            if element == "HRC":  # Vérifier si l'élément actuel est "HRC"
                valeurs_suivantes = contenu[index + 2: index + 5]  # Prendre les 3 éléments suivants
                
                # Compléter si moins de 3 valeurs sont disponibles
                while len(valeurs_suivantes) < 3:
                    valeurs_suivantes.append("Pas de valeur disponible")
                


                
                resultats["HRC"].append(
                     valeurs_suivantes
                )
            if element == "HRC":  # Vérifier si l'élément actuel est "HRC"
                valeurs_suivantes = contenu[index + 2: index + 5]  # Prendre les 3 éléments suivants
                
                # Compléter si moins de 3 valeurs sont disponibles
                while len(valeurs_suivantes) < 3:
                    valeurs_suivantes.append("Pas de valeur disponible")
                


                
                resultats["HRC"].append(
                     [""]
                )
                
                
                
            if "HRC" in resultats and isinstance(resultats["HRC"], list) and resultats["HRC"]:
                moyennes_hrc = []
                valeurs_hrc_separees = []  # Stocker les valeurs HRC séparées

                for liste_hrc in resultats["HRC"]:
                    if isinstance(liste_hrc, str) and liste_hrc.strip():  
                        liste_hrc = liste_hrc.split(", ")

                    try:
                        valeurs_hrc = [int(v.strip()) for v in liste_hrc[:3] if v.strip().replace(".", "").isdigit()]
                    except ValueError:
                        valeurs_hrc = []

                    valeurs_hrc_separees.extend(valeurs_hrc + [""])  # Ajoute une case vide après chaque groupe

                    if valeurs_hrc:
                        moyenne_hrc = round(np.mean(valeurs_hrc))
                        moyennes_hrc.append(moyenne_hrc)
                        moyennes_hrc.append("")  # Case vide après chaque moyenne

                resultats["Moyenne_HRC"] = moyennes_hrc  
                resultats["Valeurs_HRC"] = valeurs_hrc_separees  # Stocker les valeurs HRC séparées

                print("Valeurs HRC séparées avant l'export :", resultats["Valeurs_HRC"])
                print("Moyenne_HRC avant l'export :", resultats["Moyenne_HRC"])


   

                
    # Afficher les résultats après le parcours complet
    print("Résultats des informations détectées :")
    for terme, occurrences in resultats.items():
        print(f"\nTerme : {terme}")
        if occurrences:
            for occurrence in occurrences:
                print(f" - {occurrence}")
        else:
            print(" - Aucun résultat trouvé")
           
            
            
    def nettoyer_resultats(resultats):
        for key, valeurs in resultats.items():
            resultats[key] = [re.sub(r"[{}'\[\]]", "", str(val)) for val in valeurs]
        return resultats
           
    def formater_excel(chemin_fichier, colonnes_cc, colonnes_special_test):
   
    
    # Charger le fichier Excel
        wb = openpyxl.load_workbook(chemin_fichier)
        ws = wb.active

    # Vérification si la feuille est bien chargée
        if ws is None:
            print("Erreur : Impossible de charger la feuille Excel.")
            return

    # Déterminer les colonnes correspondantes en Excel
        col_cc_debut = "A"
        col_cc_fin = openpyxl.utils.get_column_letter(len(colonnes_cc))
        col_special_test_debut = openpyxl.utils.get_column_letter(len(colonnes_cc) + 1)
        col_special_test_fin = openpyxl.utils.get_column_letter(len(colonnes_cc) + len(colonnes_special_test))

    # Insérer une ligne pour les en-têtes
        ws.insert_rows(1)

    # Assigner la valeur AVANT de fusionner
        ws[f"{col_cc_debut}1"].value = "Curve"
        ws[f"{col_special_test_debut}1"].value = "Special Test"

    # Fusionner les cellules
        ws.merge_cells(f"{col_cc_debut}1:{col_cc_fin}1")
        ws.merge_cells(f"{col_special_test_debut}1:{col_special_test_fin}1")

    # Appliquer l'alignement centré aux en-têtes fusionnées
        ws[f"{col_cc_debut}1"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"{col_special_test_debut}1"].alignment = Alignment(horizontal="center", vertical="center")

    # Insérer les noms des colonnes sous "Curve" et "Special Test"
        for i, col in enumerate(colonnes_cc + colonnes_special_test):
            cell_ref = f"{openpyxl.utils.get_column_letter(i+1)}2"
            ws[cell_ref].value = col
            ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center")
        
        # for row in range(ws.max_row, 1, -1):  # On parcourt les lignes en sens inverse
        #     cell_value = resultats["IDENTITY"]  
        #     if cell_value == "IDENTITY":
        #         ws.insert_rows(row + 1)  # Insérer une ligne juste après "IDENTITY"

    # Couleurs bleu, blanc et rouge
        fill_bleu_accent = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")  
        fill_bleu_clair = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid") 
        fill_gris_clair = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  
        fill_rouge_clair = PatternFill(start_color="F4A8A8", end_color="F4A8A8", fill_type="solid")  

        for row in range(1, 3):  # Lignes 1 et 2
            for i in range(1, ws.max_column + 1):
                ws[f"{openpyxl.utils.get_column_letter(i)}{row}"].fill = fill_bleu_accent

# Alternance de couleurs dans les colonnes sous "Curve"
        for row in range(3, ws.max_row + 1):
            fill_color = [fill_bleu_clair, fill_gris_clair, fill_rouge_clair][((row - 3) // 2) % 3]  # Alterne toutes les 2 lignes
            for i, col in enumerate(colonnes_cc):
                ws[f"{openpyxl.utils.get_column_letter(i+1)}{row}"].fill = fill_color

        ws[f"{col_cc_debut}1"].fill = fill_bleu_clair  # Couleur pour l'en-tête

# Alternance des couleurs par ligne sous "Special Test"
        for row in range(3, ws.max_row + 1):
            fill_color = [fill_bleu_clair, fill_gris_clair, fill_rouge_clair][((row  - 3) // 2) % 3]
            for i, col in enumerate(colonnes_special_test, start=len(colonnes_cc) + 1):
                ws[f"{openpyxl.utils.get_column_letter(i)}{row}"].fill = fill_color

        ws[f"{col_special_test_debut}1"].fill = fill_bleu_clair  # Couleur pour l'en-tête



        # Définition des styles de bordure
        bordure = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
            )

    # Appliquer les bordures aux cellules sous "Curve"
        for i, col in enumerate(colonnes_cc):
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{openpyxl.utils.get_column_letter(i+1)}{row}"]
                cell.border = bordure  # Ajout de la bordure

# Appliquer les bordures aux cellules sous "Special Test"
        for i, col in enumerate(colonnes_special_test, start=len(colonnes_cc) + 1):
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{openpyxl.utils.get_column_letter(i)}{row}"]
                cell.border = bordure  # Ajout de la bordure

        for i, col in enumerate(colonnes_cc + colonnes_special_test):
            cell_ref = f"{openpyxl.utils.get_column_letter(i+1)}2"
            ws[cell_ref].font = Font(bold=False)

# Fusionner chaque cellule contenant un terme avec sa valeur
        for row in range(2, ws.max_row + 1):  # Commence à la ligne 2 pour éviter l'entête
            for i, col in enumerate(colonnes_cc + colonnes_special_test):
                col_letter = openpyxl.utils.get_column_letter(i+1)
                ws.merge_cells(f"{col_letter}{row}:{col_letter}{row}")  # Fusionne la case avec elle-même
                ws[f"{col_letter}{row}"].alignment = Alignment(horizontal="center", vertical="center")  # Centrage du texte        
        # Sauvegarder le fichier avec les modifications
        wb.save(chemin_fichier)
        print(" Le fichier 'resultats.xlsx' a été formaté avec succès !")   
   
    resultats = nettoyer_resultats(resultats)
    
    # Uniformisation de la longueur des listes
    max_length = max(len(v) for v in resultats.values())

    for key in resultats:
        while len(resultats[key]) < max_length:
            resultats[key].append("")  # Ajout de cellules vides
     
    chemin_fichier = r"C:\Abaqus_temp\resultats.xlsx"
    colonnes_cc = ["BAR","DIAMETER", "Elong.4D", "Elong.5D", "InitialD", "Proof(0.2%)", "mE"]
    colonnes_special_test = ["RT UTS","450°C UTS",  "RT 0.2%Proof","450°C 0.2%Proof","ElongatFracture", "ElongafterFracture", "HRC","Moyenne_HRC"]
    
    # Création du DataFrame avec les groupes de colonnes
    df = pd.DataFrame({col: resultats[col] for col in colonnes_cc + colonnes_special_test})

    df.to_excel(chemin_fichier, index=False, sheet_name="Sheet1")
        
    def sauvegarder_resultats_excel(resultats, chemin_fichier):
        try:
            df_ancien = pd.read_excel(chemin_fichier, sheet_name="Sheet1", engine="openpyxl")
        except FileNotFoundError:
                df_ancien = pd.DataFrame()

    # Convertir les nouveaux résultats en DataFrame
        df_nouveaux = pd.DataFrame(resultats)

   
        # Convertir les nouveaux résultats en DataFrame
        df_nouveaux = pd.DataFrame(resultats)

# Vérifier les colonnes
        colonnes_existantes = df_ancien.columns.tolist() if not df_ancien.empty else df_nouveaux.columns.tolist()
        df_nouveaux = df_nouveaux.reindex(columns=colonnes_existantes, fill_value="")


        df_nouveaux = df_nouveaux[~df_nouveaux.apply(tuple, axis=1).isin(df_ancien.apply(tuple, axis=1))]

        if df_nouveaux.empty:
            print(" Aucune nouvelle donnée à ajouter !")
            return


        df_total = pd.concat([df_ancien, df_nouveaux])


        with pd.ExcelWriter(chemin_fichier, engine="openpyxl", mode="w") as writer:
            df_total.to_excel(writer, sheet_name="Sheet1", index=False)

        print(f" Nouvelles données ajoutées sans duplication à '{chemin_fichier}' !")

    formater_excel(chemin_fichier, colonnes_cc, colonnes_special_test)
   
    return resultats

    
if __name__ == "__main__":
    main()

    

    













    


    