import streamlit as st
import fitz  # PyMuPDF
import pandas as pd
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment

st.set_page_config(page_title="Convertisseur PDF → Excel", layout="centered")
st.title("📄 Convertisseur PDF → Excel (Format Résultat 4)")
st.markdown("Dépose ton fichier PDF ci-dessous. Les données seront extraites et converties en Excel avec le format exact de `resultats 4.xlsx`.")

def extract_data_from_pdf(pdf_bytes):
    data = {
        "BAR": [], "DIAMETER": [], "Elong.4D": [], "Elong.5D": [], "InitialD": [],
        "Proof(0.2%)": [], "mE": [], "RT UTS": [], "450°C UTS": [], "RT 0.2%Proof": [],
        "450°C 0.2%Proof": [], "ElongatFracture": [], "ElongafterFracture": [],
        "HRC": [], "Moyenne_HRC": []
    }

    with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
        full_text = "\n".join([page.get_text() for page in doc])

    bar_match = re.search(r"CAST\*[\s\n]+([A-Z0-9]+)[\s\n]+Serial No\. ([0-9/]+)", full_text)
    if bar_match:
        data["BAR"] = [bar_match.group(1), ""]
        data["DIAMETER"] = [bar_match.group(2), ""]

    data["RT UTS"] = re.findall(r"RT.*?UTS.*?≥ \d+\n(\d+)", full_text)[:2]
    data["450°C UTS"] = re.findall(r"450°C.*?UTS.*?≥ \d+\n([\d.]+)", full_text)[:2]
    data["RT 0.2%Proof"] = re.findall(r"RT.*?0\.2% Proof.*?≥ \d+\n(\d+)", full_text)[:2]
    data["450°C 0.2%Proof"] = re.findall(r"450°C.*?0\.2% Proof.*?≥ \d+\n([\d.]+)", full_text)[:2]

    for key in data:
        while len(data[key]) < 2:
            data[key].append("")

    return data

def create_excel(data_dict):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active

    ws.merge_cells("A1:G1")
    ws["A1"] = "Curve"
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("H1:O1")
    ws["H1"] = "Special Test"
    ws["H1"].alignment = Alignment(horizontal="center", vertical="center")

    headers = list(data_dict.keys())
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx in range(2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx + 3, column=col_idx).value = data_dict[key][row_idx]

    wb.save(output)
    output.seek(0)
    return output

uploaded_file = st.file_uploader("Choisis un fichier PDF", type="pdf")

if uploaded_file:
    with st.spinner("📤 Traitement du fichier..."):
        pdf_bytes = uploaded_file.read()
        extracted_data = extract_data_from_pdf(pdf_bytes)
        excel_file = create_excel(extracted_data)

        st.success("✅ Fichier traité avec succès !")
        st.download_button(
            label="📥 Télécharger le fichier Excel",
            data=excel_file,
            file_name="resultats_formaté.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
