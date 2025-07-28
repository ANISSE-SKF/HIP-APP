import streamlit as st
import fitz  # PyMuPDF
import pandas as pd
import io
import re
import openpyxl
import gc
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
import numpy as np

st.set_page_config(page_title="Convertisseur PDF → Excel", layout="centered")
st.title("📄 Convertisseur PDF → Excel (Format Résultat)")
st.markdown("Dépose ton fichier PDF ci-dessous. Les données seront extraites et converties en Excel avec le format attendu.")

def extract_text_from_first_page(pdf_bytes):
    with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
        if len(doc) > 0:
            page = doc[0]
            return page.get_text()
    return ""

def parse_text_to_list(text):
    lines = text.splitlines()
    return [line.split() for line in lines if line.strip()]

def create_dictionaries_from_lists(listes):
    dictionnaire = {}
    for i, liste in enumerate(listes):
        dictionnaire[f"Liste_{i+1}"] = liste
    return dictionnaire

def nettoyer_dictionnaire(resultats):
    dictionnaire_nettoye = {}
    for cle, liste in resultats.items():
        dictionnaire_nettoye[cle] = [str(element).replace("'", "").replace(",", "").replace(" ", "") for element in liste if element]
    return dictionnaire_nettoye

def detecter_et_traiter_informations(dictionnaire):
    resultats = {"Elong.4D": [], "Elong.5D": [], "InitialD": [], "Proof(0.2%)": [], "mE": [],
                 "RT UTS": [], "450°C UTS": [], "BAR": [], "DIAMETER": [], "ElongatFracture": [],
                 "ElongafterFracture": [], "RT 0.2%Proof": [], "450°C 0.2%Proof": [], "HRC": [], "Moyenne_HRC": []}
    # Pour simplifier ici, on extrait les valeurs connues depuis le texte brut
    texte = "\n".join([" ".join(ligne) for ligne in dictionnaire.values()])
    bar_match = re.search(r"CAST\*[\s\n]+([A-Z0-9]+)[\s\n]+Serial No\. ([0-9/]+)", texte)
    if bar_match:
        resultats["BAR"] = [bar_match.group(1), ""]
        resultats["DIAMETER"] = [bar_match.group(2), ""]

    resultats["RT UTS"] = re.findall(r"RT.*?UTS.*?≥ \d+\n?(\d+)", texte)[:2]
    resultats["450°C UTS"] = re.findall(r"450°C.*?UTS.*?≥ \d+\n?([\d.]+)", texte)[:2]
    resultats["RT 0.2%Proof"] = re.findall(r"RT.*?0\.2% Proof.*?≥ \d+\n?(\d+)", texte)[:2]
    resultats["450°C 0.2%Proof"] = re.findall(r"450°C.*?0\.2% Proof.*?≥ \d+\n?([\d.]+)", texte)[:2]
    resultats["ElongatFracture"] = re.findall(r"RT.*?Elong at Fracture.*?(\d+)%", texte)[:2]
    resultats["ElongafterFracture"] = re.findall(r"450°C.*?Elong after Fracture.*?(\d+\.?\d*)%", texte)[:2]

    hrc_values = re.findall(r"HRC.*?\n(\d+)\n(\d+)\n(\d+)", texte)
    if hrc_values:
        hrc = hrc_values[0]
        resultats["HRC"] = [", ".join(hrc), ""]
        moyenne = round(sum(map(int, hrc)) / 3)
        resultats["Moyenne_HRC"] = [str(moyenne), ""]

    for key in resultats:
        while len(resultats[key]) < 2:
            resultats[key].append("")

    return resultats

def create_excel(resultats):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active

    colonnes_cc = ["BAR", "DIAMETER", "Elong.4D", "Elong.5D", "InitialD", "Proof(0.2%)", "mE"]
    colonnes_special_test = ["RT UTS", "450°C UTS", "RT 0.2%Proof", "450°C 0.2%Proof",
                             "ElongatFracture", "ElongafterFracture", "HRC", "Moyenne_HRC"]

    ws.merge_cells("A1:G1")
    ws["A1"] = "Curve"
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("H1:O1")
    ws["H1"] = "Special Test"
    ws["H1"].alignment = Alignment(horizontal="center", vertical="center")

    headers = colonnes_cc + colonnes_special_test
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx in range(2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx + 3, column=col_idx).value = resultats[key][row_idx]

    wb.save(output)
    output.seek(0)
    return output

uploaded_file = st.file_uploader("Dépose ton fichier PDF ici", type="pdf")

if uploaded_file:
    with st.spinner("📤 Traitement du fichier..."):
        pdf_bytes = uploaded_file.read()
        text = extract_text_from_first_page(pdf_bytes)
        lignes = parse_text_to_list(text)
        dictionnaire = create_dictionaries_from_lists(lignes)
        dictionnaire_nettoye = nettoyer_dictionnaire(dictionnaire)
        resultats = detecter_et_traiter_informations(dictionnaire_nettoye)
        excel_file = create_excel(resultats)

        st.success("✅ Fichier traité avec succès !")
        st.download_button(
            label="📥 Télécharger le fichier Excel",
            data=excel_file,
            file_name="resultats.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


